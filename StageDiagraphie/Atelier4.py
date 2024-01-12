import numpy as np
from openpyxl import load_workbook
import matplotlib.pyplot as plt 

wb = load_workbook("/Users/merliemmanuelendamneobiang/Downloads/sws_aubure_X1971_220524130801.xlsx")

ws = wb.active

pressure = ws["B"]
temperature = ws["C"]
conductivity = ws["D"]

tableauDePression = np.zeros(len(pressure))
tableauDeTemperature = np.zeros(len(temperature))
tableauDeConductivite = np.zeros(len(conductivity))



for k in range(1,len(pressure)) :
    tableauDePression[k] = float(pressure[k].value)

    tableauDeTemperature[k] = float(temperature[k].value)
    
    tableauDeConductivite[k] = float(conductivity[k].value)
    
fig, (ax1, ax2, ax3) = plt.subplots( nrows = 3, ncols = 1)

ax1.plot(tableauDePression, "k")
ax1.set_title("Pression")
ax2.plot(tableauDeTemperature, "b")
ax2.set_title("Température")
ax2.set_xlim(700,1400)
ax2.set_ylim(5,11)
ax3.plot(tableauDeConductivite, "r")
ax3.set_title("Conductivité")

plt.tight_layout()
plt.show()  
  
    
